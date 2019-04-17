
# openpyxl

import openpyxl

planilha = openpyxl.Workbook()

#print(planilha.sheetnames)  # retorna abas

planilha.create_sheet(index=0, title='Laticinios')
planilha.create_sheet(index=1, title='Frutas')
planilha.create_sheet(index=2, title='Verduras')
planilha.create_sheet(index=3, title='Padaria')

#print(planilha.sheetnames)

del planilha['Sheet']

print(planilha.sheetnames)

planilha.save('mercado.xlsx')
