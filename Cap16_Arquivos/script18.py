

from openpyxl import load_workbook

planilha = load_workbook('mercado.xlsx')

print(planilha.sheetnames)

l = planilha['Laticinios']
l['A1'] = 'Produto'
l['A2'] = 'Nan'
l['A3'] = 'Molico'
l['A4'] = 'Ninho'
l['A5'] = 'Nescau'

l['B1'] = 'Valor'

l['C1'] = 'Qtd'


planilha.save('mercado2.xlsx')
